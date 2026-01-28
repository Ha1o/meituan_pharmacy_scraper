"""
生成示例任务文件 tasks_template.xlsx
运行此脚本生成模板文件
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os


def create_template():
    """创建示例任务文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "采集任务"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ["poi", "shop_name", "note"]
    header_widths = [30, 40, 20]
    
    for col, (header, width) in enumerate(zip(headers, header_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col)].width = width
    
    # 示例数据
    sample_data = [
        ("天河区体育西路", "好药师大药房（体育西店）", "示例任务1"),
        ("海珠区江南西", "大参林（江南西店）", "示例任务2"),
        ("越秀区北京路", "老百姓大药房（北京路店）", "示例任务3"),
        ("番禺区市桥", "海王星辰（市桥店）", ""),
        ("白云区白云新城", "一心堂（白云新城店）", ""),
    ]
    
    # 数据样式
    data_alignment = Alignment(horizontal="left", vertical="center")
    
    for row_num, (poi, shop_name, note) in enumerate(sample_data, 2):
        ws.cell(row=row_num, column=1, value=poi).alignment = data_alignment
        ws.cell(row=row_num, column=1).border = thin_border
        
        ws.cell(row=row_num, column=2, value=shop_name).alignment = data_alignment
        ws.cell(row=row_num, column=2).border = thin_border
        
        ws.cell(row=row_num, column=3, value=note).alignment = data_alignment
        ws.cell(row=row_num, column=3).border = thin_border
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存
    output_dir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(output_dir, "tasks_template.xlsx")
    wb.save(filepath)
    wb.close()
    
    print(f"示例任务文件已创建: {filepath}")
    return filepath


if __name__ == "__main__":
    create_template()
