"""
task_loader.py - xlsx 任务加载器
读取任务文件，解析 poi, shop_name, note 字段
"""
import os
from typing import List, Dict, Optional
from dataclasses import dataclass
from openpyxl import load_workbook

from core.logger import DeviceLogger


@dataclass
class Task:
    """单个任务（一个店铺）"""
    index: int          # 任务序号（从0开始）
    poi: str            # 定位点关键词
    shop_name: str      # 店铺名
    note: str           # 备注（可选）
    
    def __str__(self):
        return f"任务[{self.index}]: POI={self.poi}, 店铺={self.shop_name}"


class TaskLoader:
    """
    任务加载器
    从xlsx文件读取任务列表
    """
    
    # 支持的列名映射（中文 -> 英文）
    COLUMN_MAPPING = {
        '定位点': 'poi',
        'poi': 'poi',
        '店铺名字': 'shop_name',
        '店铺名': 'shop_name',
        'shop_name': 'shop_name',
        '备注': 'note',
        'note': 'note',
        '定位id': 'location_id',  # 可选
    }
    
    # 必需的字段（英文key）
    REQUIRED_FIELDS = ['poi', 'shop_name']
    
    def __init__(self, logger: Optional[DeviceLogger] = None):
        """
        初始化任务加载器
        
        Args:
            logger: 日志器（可选）
        """
        self.logger = logger
        self.tasks: List[Task] = []
        self.file_path: str = ""
    
    def _log(self, message: str, level: str = "info"):
        """记录日志"""
        if self.logger:
            getattr(self.logger, level)(message)
    
    def load(self, file_path: str) -> bool:
        """
        加载xlsx任务文件
        
        Args:
            file_path: xlsx文件路径
            
        Returns:
            是否加载成功
        """
        self.tasks = []
        self.file_path = file_path
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            self._log(f"任务文件不存在: {file_path}", "error")
            return False
        
        # 检查文件扩展名
        if not file_path.lower().endswith('.xlsx'):
            self._log(f"文件格式错误，必须是xlsx格式: {file_path}", "error")
            return False
        
        try:
            self._log(f"正在加载任务文件: {file_path}")
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            if ws is None:
                self._log("xlsx文件中没有工作表", "error")
                return False
            
            # 读取表头（第一行）
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not header_row:
                self._log("xlsx文件表头为空", "error")
                return False
            
            # 转换表头：保留原始值（用于匹配中文）
            headers_raw = [str(h).strip() if h else "" for h in header_row]
            
            # 建立列索引映射（原始列索引 -> 英文字段名）
            col_indices = {}
            for col_idx, header in enumerate(headers_raw):
                # 尝试匹配列名
                header_lower = header.lower()
                if header_lower in self.COLUMN_MAPPING:
                    field_name = self.COLUMN_MAPPING[header_lower]
                    col_indices[field_name] = col_idx
                elif header in self.COLUMN_MAPPING:
                    field_name = self.COLUMN_MAPPING[header]
                    col_indices[field_name] = col_idx
            
            # 检查必需字段
            for field_name in self.REQUIRED_FIELDS:
                if field_name not in col_indices:
                    self._log(f"缺少必需列: {field_name}（支持的列名: 定位点/poi, 店铺名字/店铺名/shop_name）", "error")
                    return False
            
            # 读取数据行（从第2行开始）
            task_index = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 跳过空行
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                
                # 获取各列值
                poi = ""
                shop_name = ""
                note = ""
                
                if 'poi' in col_indices and col_indices['poi'] < len(row):
                    poi = str(row[col_indices['poi']]).strip() if row[col_indices['poi']] else ""
                if 'shop_name' in col_indices and col_indices['shop_name'] < len(row):
                    shop_name = str(row[col_indices['shop_name']]).strip() if row[col_indices['shop_name']] else ""
                if 'note' in col_indices and col_indices['note'] < len(row):
                    note = str(row[col_indices['note']]).strip() if row[col_indices['note']] else ""
                
                # 验证必需字段
                if not poi or not shop_name:
                    self._log(f"第{task_index + 2}行数据不完整，跳过: poi={poi}, shop_name={shop_name}", "warning")
                    continue
                
                # 创建任务
                task = Task(
                    index=task_index,
                    poi=poi,
                    shop_name=shop_name,
                    note=note
                )
                self.tasks.append(task)
                task_index += 1
            
            wb.close()
            
            self._log(f"成功加载 {len(self.tasks)} 个任务")
            return True
            
        except Exception as e:
            self._log(f"加载任务文件失败: {e}", "error")
            return False
    
    def get_tasks(self) -> List[Task]:
        """获取所有任务"""
        return self.tasks
    
    def get_task(self, index: int) -> Optional[Task]:
        """
        获取指定索引的任务
        
        Args:
            index: 任务索引
            
        Returns:
            任务对象，不存在返回None
        """
        if 0 <= index < len(self.tasks):
            return self.tasks[index]
        return None
    
    def count(self) -> int:
        """获取任务数量"""
        return len(self.tasks)
    
    def __len__(self):
        return len(self.tasks)
    
    def __iter__(self):
        return iter(self.tasks)
