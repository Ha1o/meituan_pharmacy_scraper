"""
exporter.py - Excel 导出模块
以店铺名生成xlsx文件，按需求模板格式输出
"""
import os
import re
from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from core.logger import DeviceLogger


class DrugRecord:
    """药品记录"""
    
    def __init__(
        self,
        category_name: str,
        drug_name: str,
        monthly_sales: str,
        price: str
    ):
        self.category_name = category_name
        self.drug_name = drug_name
        self.monthly_sales = monthly_sales
        self.price = price
    
    def to_dict(self) -> Dict[str, str]:
        return {
            "商品分类": self.category_name,
            "商品名字": self.drug_name,
            "月销量": self.monthly_sales,
            "价格": self.price
        }
    
    def to_list(self) -> List[str]:
        return [
            self.category_name,
            self.drug_name,
            self.monthly_sales,
            self.price
        ]


class ExcelExporter:
    """
    Excel导出器
    将采集的药品数据导出为xlsx文件
    按需求模板格式：定位ID、定位点、店铺名字、商品分类、商品名字、月销量、价格
    """
    
    # 表头（按需求模板）
    HEADERS = ["定位ID", "定位点", "店铺名字", "商品分类", "商品名字", "月销量", "价格"]
    
    def __init__(self, device_serial: str, base_output_dir: str = "output", logger: Optional[DeviceLogger] = None):
        """
        初始化导出器
        
        Args:
            device_serial: 设备序列号
            base_output_dir: 输出根目录（如 "output"）
            logger: 日志器
        """
        self.device_serial = device_serial
        self.base_output_dir = base_output_dir
        self.logger = logger
        
        # 使用 paths 模块创建目录: output/{serial}/results
        from core import paths
        self.results_dir = paths.results_dir(base_output_dir, device_serial)
        
        # 当前店铺的记录列表
        self.records: List[DrugRecord] = []
        self.current_shop_name: str = ""
        self.current_poi: str = ""  # 定位点
        self.current_task_id: int = 1  # 定位ID
    
    def _log(self, message: str, level: str = "info"):
        """记录日志"""
        if self.logger:
            getattr(self.logger, level)(message)
    
    @staticmethod
    def sanitize_filename(name: str) -> str:
        """
        清理文件名中的非法字符
        
        Args:
            name: 原始文件名
            
        Returns:
            清理后的文件名
        """
        # Windows文件名非法字符
        illegal_chars = r'[<>:"/\\|?*]'
        sanitized = re.sub(illegal_chars, '_', name)
        
        # 去除首尾空格和点
        sanitized = sanitized.strip('. ')
        
        # 限制长度
        if len(sanitized) > 100:
            sanitized = sanitized[:100]
        
        # 如果为空，使用默认名
        if not sanitized:
            sanitized = "unknown_shop"
        
        return sanitized
    
    def start_shop(self, shop_name: str, poi: str = "", task_id: int = 1):
        """
        开始新店铺的记录
        
        Args:
            shop_name: 店铺名
            poi: 定位点地址
            task_id: 任务ID（定位ID）
        """
        self.records = []
        self.current_shop_name = shop_name
        self.current_poi = poi
        self.current_task_id = task_id
        self._log(f"开始记录店铺数据: {shop_name}")
    
    def add_record(self, record: DrugRecord):
        """
        添加药品记录
        
        Args:
            record: 药品记录
        """
        self.records.append(record)
    
    def add_records(self, records: List[DrugRecord]):
        """
        批量添加药品记录
        
        Args:
            records: 记录列表
        """
        self.records.extend(records)
    
    def export(self, shop_name: Optional[str] = None) -> Optional[str]:
        """
        导出当前店铺数据到xlsx
        
        Args:
            shop_name: 店铺名（可选，不传则使用start_shop时设置的名称）
            
        Returns:
            导出文件路径，失败返回None
        """
        shop_name = shop_name or self.current_shop_name
        
        if not shop_name:
            self._log("导出失败：未设置店铺名", "error")
            return None
        
        if not self.records:
            self._log(f"店铺 [{shop_name}] 无数据，跳过导出", "warning")
            return None
        
        # 生成安全的文件名: output/{serial}/results/{shop_name}_{task_id}.xlsx
        safe_name = self.sanitize_filename(shop_name)
        filename = f"{safe_name}_{self.current_task_id}.xlsx"
        filepath = os.path.join(self.results_dir, filename)
        
        try:
            self._log(f"正在导出: {filepath}")
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "药品数据"
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入表头
            for col, header in enumerate(self.HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 写入数据
            # 按需求模板：每一行都要填充定位ID、定位点、店铺名字
            for row_num, record in enumerate(self.records, 2):
                # 每一行都写入前三列
                ws.cell(row=row_num, column=1, value=self.current_task_id)
                ws.cell(row=row_num, column=2, value=self.current_poi)
                ws.cell(row=row_num, column=3, value=shop_name)

                # 商品数据从第4列开始
                data = record.to_list()  # [分类, 商品名, 月销, 价格]
                for col, value in enumerate(data, 4):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 调整列宽
            column_widths = [10, 35, 25, 15, 40, 12, 12]
            for col, width in enumerate(column_widths, 1):
                if col <= len(column_widths):
                    ws.column_dimensions[chr(64 + col)].width = width
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存
            wb.save(filepath)
            wb.close()
            
            self._log(f"导出成功: {filepath} (共{len(self.records)}条记录)")
            return filepath
            
        except Exception as e:
            self._log(f"导出失败: {e}", "error")
            return None
    
    def get_record_count(self) -> int:
        """获取当前记录数"""
        return len(self.records)
    
    def clear(self):
        """清空记录"""
        self.records = []
        self.current_shop_name = ""
        self.current_poi = ""
        self.current_task_id = 1


def create_drug_record(
    category_name: str,
    drug_name: str,
    monthly_sales: str = "0",
    price: str = ""
) -> DrugRecord:
    """
    创建药品记录的便捷函数
    
    Args:
        category_name: 分类名
        drug_name: 药品名
        monthly_sales: 月销（默认0）
        price: 价格
        
    Returns:
        DrugRecord对象
    """
    # 清理月销数据
    # 强制格式：月售 + 数字
    if not monthly_sales or monthly_sales.strip() == "":
        num = "0"
    else:
        # 提取数字
        sales_match = re.search(r'(\d+)', monthly_sales)
        num = sales_match.group(1) if sales_match else "0"

    monthly_sales = f"月售{num}"
    
    # 清理价格数据
    if price:
        price = price.strip().replace('¥', '').replace('￥', '')
    
    return DrugRecord(
        category_name=category_name.strip(),
        drug_name=drug_name.strip(),
        monthly_sales=monthly_sales,
        price=price
    )
